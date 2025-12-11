"""
Excel and CSV processor for bulk CV uploads
"""
import io
import csv
from typing import List, Dict, Any
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_excel_file(excel_bytes: bytes) -> List[Dict[str, Any]]:
    """
    מפרק קובץ אקסל ומחזיר רשימת רשומות
    
    Args:
        excel_bytes: תוכן הקובץ bytes
    
    Returns:
        רשימה של dictionaries, כל אחד מכיל את הנתונים של שורה אחת
        כל dictionary מכיל: name, phone_number, email, campaign, notes
    """
    try:
        # טען את הקובץ
        workbook = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        sheet = workbook.active
        
        # מצא את השורה הראשונה (headers)
        headers = []
        header_row = None
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
            # נסה למצוא שורה עם headers
            row_values = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
            
            # בדוק אם יש את העמודות הנדרשות
            required_columns = ["name", "phone_number", "phone", "email", "campaign", "notes"]
            found_columns = []
            column_indices = {}
            
            for col_idx, header_value in enumerate(row_values, start=1):
                header_lower = header_value.lower().strip()
                # נסה להתאים לעמודות אפשריות
                if "name" in header_lower:
                    column_indices["name"] = col_idx
                    found_columns.append("name")
                elif "phone" in header_lower or "tel" in header_lower:
                    column_indices["phone_number"] = col_idx
                    found_columns.append("phone_number")
                elif "email" in header_lower or "mail" in header_lower:
                    column_indices["email"] = col_idx
                    found_columns.append("email")
                elif "campaign" in header_lower:
                    column_indices["campaign"] = col_idx
                    found_columns.append("campaign")
                elif "notes" in header_lower or "note" in header_lower:
                    column_indices["notes"] = col_idx
                    found_columns.append("notes")
            
            # אם מצאנו לפחות כמה עמודות, זה כנראה שורת headers
            if len(found_columns) >= 2:  # לפחות name ו-phone
                header_row = row_idx
                headers = row_values
                break
        
        if header_row is None:
            raise ValueError("לא נמצאו headers בקובץ האקסל. יש לוודא שיש שורת headers עם העמודות: name, phone_number, email, campaign, notes")
        
        # קרא את הנתונים
        records = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            # דלג על שורות ריקות
            if not any(cell for cell in row if cell):
                continue
            
            # חלץ את הנתונים לפי העמודות
            record = {}
            
            # name
            if "name" in column_indices:
                name_idx = column_indices["name"] - 1
                record["name"] = str(row[name_idx]).strip() if row[name_idx] else None
            else:
                record["name"] = None
            
            # phone_number
            if "phone_number" in column_indices:
                phone_idx = column_indices["phone_number"] - 1
                phone_value = row[phone_idx] if phone_idx < len(row) else None
                # המר מספר לטלפון ל-string
                if phone_value is not None:
                    record["phone_number"] = str(phone_value).strip()
                else:
                    record["phone_number"] = None
            else:
                record["phone_number"] = None
            
            # email
            if "email" in column_indices:
                email_idx = column_indices["email"] - 1
                record["email"] = str(row[email_idx]).strip() if email_idx < len(row) and row[email_idx] else None
            else:
                record["email"] = None
            
            # campaign
            if "campaign" in column_indices:
                campaign_idx = column_indices["campaign"] - 1
                record["campaign"] = str(row[campaign_idx]).strip() if campaign_idx < len(row) and row[campaign_idx] else None
            else:
                record["campaign"] = None
            
            # notes
            if "notes" in column_indices:
                notes_idx = column_indices["notes"] - 1
                record["notes"] = str(row[notes_idx]).strip() if notes_idx < len(row) and row[notes_idx] else None
            else:
                record["notes"] = None
            
            # דלג על רשומות ריקות לחלוטין
            if not any([record.get("name"), record.get("phone_number"), record.get("email")]):
                continue
            
            records.append(record)
        
        logger.info(f"[EXCEL] Parsed {len(records)} records from Excel file")
        return records
        
    except Exception as e:
        logger.error(f"[EXCEL] Error parsing Excel file: {str(e)}", exc_info=True)
        raise ValueError(f"שגיאה בפרסור קובץ האקסל: {str(e)}")


def parse_csv_file(csv_bytes: bytes, encoding: str = 'utf-8') -> List[Dict[str, Any]]:
    """
    מפרק קובץ CSV ומחזיר רשימת רשומות
    
    Args:
        csv_bytes: תוכן הקובץ bytes
        encoding: קידוד הקובץ (ברירת מחדל: utf-8)
    
    Returns:
        רשימה של dictionaries, כל אחד מכיל את הנתונים של שורה אחת
        כל dictionary מכיל: name, phone_number, email, campaign, notes
    """
    try:
        # נסה לקרוא עם encoding שונים
        encodings_to_try = [encoding, 'utf-8-sig', 'latin-1', 'iso-8859-1']
        csv_text = None
        
        for enc in encodings_to_try:
            try:
                csv_text = csv_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        
        if csv_text is None:
            raise ValueError("לא ניתן לפענח את קובץ ה-CSV. נסה לשמור את הקובץ ב-UTF-8")
        
        # קרא את הקובץ
        csv_reader = csv.DictReader(io.StringIO(csv_text))
        
        # מצא את העמודות
        column_mapping = {}
        headers = csv_reader.fieldnames
        
        if not headers:
            raise ValueError("לא נמצאו headers בקובץ ה-CSV")
        
        # התאם את העמודות
        for header in headers:
            header_lower = header.lower().strip()
            if "name" in header_lower:
                column_mapping["name"] = header
            elif "phone" in header_lower or "tel" in header_lower:
                column_mapping["phone_number"] = header
            elif "email" in header_lower or "mail" in header_lower:
                column_mapping["email"] = header
            elif "campaign" in header_lower:
                column_mapping["campaign"] = header
            elif "notes" in header_lower or "note" in header_lower:
                column_mapping["notes"] = header
        
        # בדוק שיש לפחות name או phone_number
        if "name" not in column_mapping and "phone_number" not in column_mapping:
            raise ValueError("לא נמצאו עמודות name או phone_number בקובץ ה-CSV")
        
        # קרא את הנתונים
        records = []
        for row_idx, row in enumerate(csv_reader, start=2):  # start=2 כי שורה 1 היא headers
            # דלג על שורות ריקות
            if not any(row.values()):
                continue
            
            # חלץ את הנתונים לפי העמודות
            record = {}
            
            # name
            if "name" in column_mapping:
                record["name"] = row.get(column_mapping["name"], "").strip() or None
            else:
                record["name"] = None
            
            # phone_number
            if "phone_number" in column_mapping:
                phone_value = row.get(column_mapping["phone_number"], "").strip()
                record["phone_number"] = phone_value if phone_value else None
            else:
                record["phone_number"] = None
            
            # email
            if "email" in column_mapping:
                record["email"] = row.get(column_mapping["email"], "").strip() or None
            else:
                record["email"] = None
            
            # campaign
            if "campaign" in column_mapping:
                record["campaign"] = row.get(column_mapping["campaign"], "").strip() or None
            else:
                record["campaign"] = None
            
            # notes
            if "notes" in column_mapping:
                record["notes"] = row.get(column_mapping["notes"], "").strip() or None
            else:
                record["notes"] = None
            
            # דלג על רשומות ריקות לחלוטין
            if not any([record.get("name"), record.get("phone_number"), record.get("email")]):
                continue
            
            records.append(record)
        
        logger.info(f"[CSV] Parsed {len(records)} records from CSV file")
        return records
        
    except Exception as e:
        logger.error(f"[CSV] Error parsing CSV file: {str(e)}", exc_info=True)
        raise ValueError(f"שגיאה בפרסור קובץ ה-CSV: {str(e)}")

